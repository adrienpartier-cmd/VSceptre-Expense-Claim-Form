from llama_cloud_services import LlamaExtract
from pydantic import BaseModel, Field, ValidationError
from typing import Optional
from pathlib import Path
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.worksheet.pagebreak import Break
from openpyxl.styles import Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import range_boundaries
from copy import copy, deepcopy
from datetime import datetime
from PIL import Image as PILImage
import math
import os
import shutil
import textwrap
import warnings

warnings.filterwarnings(
    "ignore",
    message="Data Validation extension is not supported and will be removed"
)

BASE_DIR = Path(__file__).resolve().parent
MASTER_FILE = BASE_DIR / "Expense_Claim_MASTER.xlsx"
OUTPUT_FILE = BASE_DIR / "Expense_Claim_filled.xlsx"
RECEIPTS_FOLDER = BASE_DIR / "receipts"

SHEET_NAME = "Form"
DROPDOWN_SHEET = "Dropdown"

HEADER_ROW = 6
RECEIPT_TEMPLATE_ROW = 7
SUMMARY_HEADER_TEMPLATE_ROW = 8
CATEGORY_TEMPLATE_ROW = 9
GRAND_TOTAL_TEMPLATE_ROW = 10
FOOTER_FIRST_ROW = 11
FOOTER_LAST_ROW = 24

ITEM_COLUMN = "A"
DATE_COLUMN = "B"
MERCHANT_COLUMN = "C"
DESCRIPTION_COLUMN = "D"
ADDITIONAL_CONTEXT_START_COLUMN = "E"
ADDITIONAL_CONTEXT_END_COLUMN = "F"
CATEGORY_COLUMN = "G"
AMOUNT_COLUMN = "H"
AMOUNT_END_COLUMN = "I"

load_dotenv(BASE_DIR / ".env")
API_KEY = os.getenv("LLAMA_CLOUD_API_KEY")

ALLOWED_CATEGORIES = [
    "Office/Computer Equipment",
    "Furniture & Fixtures",
    "Purchase",
    "Software - Purchase",
    "Misc. Parts - Purchase",
    "Audit Fee",
    "Bank Charges",
    "Business Registration Fee",
    "Consulting and Secretarial Fee",
    "Legal and Professional Fee",
    "Finance Costs",
    "General Office Expense",
    "Insurance - Office",
    "Insurance -Staff (non-sales)",
    "Insurance - Contractor",
    "Print,Stationery, Due & Subs",
    "Postage & Courier",
    "Rent & Management Fee - Office",
    "Rates & Gov Rent - Office",
    "Repairs & Maintenance - Office",
    "Telephone & Broadband - Office",
    "Utilities - Office",
    "MPF (Non-Sales)",
    "Legal Fees",
    "Office Supplies",
    "Selling & Distribution Expense",
    "Advertising &Promotion Expense",
    "Entertainment",
    "Sundry Expense- Office",
    "Travelling - Overseas",
    "Travelling - Local",
    "Training and Seminar Costs",
    "Staff Amenities",
    "Salaries & Allowances (Sales)",
    "Wages & Salaries",
    "/",
]

CATEGORY_PROMPT = """
Choose EXACTLY ONE of the following Excel dropdown options.

Return the wording EXACTLY as written.

Office/Computer Equipment
Furniture & Fixtures
Purchase
Software - Purchase
Misc. Parts - Purchase
Audit Fee
Bank Charges
Business Registration Fee
Consulting and Secretarial Fee
Legal and Professional Fee
Finance Costs
General Office Expense
Insurance - Office
Insurance -Staff (non-sales)
Insurance - Contractor
Print,Stationery, Due & Subs
Postage & Courier
Rent & Management Fee - Office
Rates & Gov Rent - Office
Repairs & Maintenance - Office
Telephone & Broadband - Office
Utilities - Office
MPF (Non-Sales)
Legal Fees
Office Supplies
Selling & Distribution Expense
Advertising &Promotion Expense
Entertainment
Sundry Expense- Office
Travelling - Overseas
Travelling - Local
Training and Seminar Costs
Staff Amenities
Salaries & Allowances (Sales)
Wages & Salaries
/

Classification guidance:

Computers, laptops, monitors, computer hardware and similar equipment
-> Office/Computer Equipment

Furniture, desks, chairs and cabinets
-> Furniture & Fixtures

General purchases which do not fit another specific category
-> Purchase

Software purchases or licences
-> Software - Purchase

Small electronic or computer replacement parts
-> Misc. Parts - Purchase

Audit charges
-> Audit Fee

Bank service or transaction charges
-> Bank Charges

Business registration payments
-> Business Registration Fee

Consulting or company-secretarial services
-> Consulting and Secretarial Fee

General professional services
-> Legal and Professional Fee

Finance or interest costs
-> Finance Costs

General routine office expenses
-> General Office Expense

Office insurance
-> Insurance - Office

Non-sales staff insurance
-> Insurance -Staff (non-sales)

Contractor insurance
-> Insurance - Contractor

Printing, stationery, subscriptions or membership dues
-> Print,Stationery, Due & Subs

Courier, delivery or postage
-> Postage & Courier

Office rental and management fees
-> Rent & Management Fee - Office

Government rates or office government rent
-> Rates & Gov Rent - Office

Repairs and maintenance
-> Repairs & Maintenance - Office

Telephone, mobile, internet or broadband
-> Telephone & Broadband - Office

Electricity, water or utilities
-> Utilities - Office

Mandatory Provident Fund
-> MPF (Non-Sales)

Legal work specifically
-> Legal Fees

Office stationery, paper, pens and ordinary office supplies
-> Office Supplies

Selling or distribution expenses
-> Selling & Distribution Expense

Advertising, promotion or marketing
-> Advertising &Promotion Expense

Restaurants, cafes, meals, drinks, food or business meals
-> Entertainment

Miscellaneous office expense that does not fit another category
-> Sundry Expense- Office

Flights and overseas transport/travel
-> Travelling - Overseas

Hong Kong taxi, Uber, MTR, bus, ferry, parking, tolls or local travel
-> Travelling - Local

Courses, seminars, conferences, examinations or certifications
-> Training and Seminar Costs

Staff refreshments, welfare or amenities
-> Staff Amenities

Sales staff salaries or allowances
-> Salaries & Allowances (Sales)

General wages and salaries
-> Wages & Salaries

If no sensible category can be determined
-> /

Do NOT create a new category.
"""


class Receipt(BaseModel):
    merchant_name: str = Field(
        description="Business, shop, restaurant or organisation name printed on the receipt."
    )
    receipt_date: str = Field(
        description="Transaction date in DD/MM/YYYY format."
    )
    description: str = Field(
        description="Concise factual description of the items or services purchased. Include important product names or model numbers visible on the receipt where useful. Do not invent information."
    )
    total_hkd: float = Field(
        description="Final amount paid as printed on the receipt. Do NOT perform currency conversion."
    )
    currency: str = Field(
        description="Three-letter currency code such as HKD, USD, GBP, EUR, JPY or TWD."
    )
    category: str = Field(
        description=CATEGORY_PROMPT
    )
    confidence_note: Optional[str] = Field(
        default=None,
        description="Briefly explain if important receipt information is unclear. Otherwise return null."
    )


def check_master_file():
    if not MASTER_FILE.exists():
        raise SystemExit(
            f"\nERROR: Expense_Claim_MASTER.xlsx was not found.\n\nExpected:\n{MASTER_FILE}\n"
        )


def check_receipts_folder():
    if not RECEIPTS_FOLDER.exists():
        RECEIPTS_FOLDER.mkdir()

        raise SystemExit(
            f"\nCreated receipts folder:\n\n{RECEIPTS_FOLDER}\n\nPut receipt files inside it and run the program again."
        )


def get_receipt_files():
    files = sorted(
        file
        for file in RECEIPTS_FOLDER.iterdir()
        if file.suffix.lower() in (
            ".jpg",
            ".jpeg",
            ".png",
            ".pdf",
        )
    )

    if not files:
        raise SystemExit(
            "\nNo receipt files found inside the receipts folder."
        )

    return files


def get_agent():
    if not API_KEY:
        raise SystemExit(
            "\nMissing LLAMA_CLOUD_API_KEY.\n\n"
            "Your API_KEY.txt file should contain:\n\n"
            "LLAMA_CLOUD_API_KEY=llx-your-real-key\n"
        )

    extractor = LlamaExtract(
        api_key=API_KEY
    )

    agent_name = "expense-dynamic-final-v6"

    try:
        agent = extractor.get_agent(
            name=agent_name
        )

        print(
            f"\nUsing extraction agent '{agent_name}'."
        )

        return agent

    except Exception:
        print(
            f"\nCreating extraction agent '{agent_name}'..."
        )

        return extractor.create_agent(
            name=agent_name,
            data_schema=Receipt
        )


def normalise_category(category):
    if not category:
        return "/"

    category = str(
        category
    ).strip()

    if category in ALLOWED_CATEGORIES:
        return category

    for allowed in ALLOWED_CATEGORIES:
        if category.casefold() == allowed.casefold():
            return allowed

    return "/"


def extract_receipts(files):
    agent = get_agent()
    extracted = []
    failed = []

    print(
        f"\nFound {len(files)} receipt(s).\n"
    )

    for number, file in enumerate(
        files,
        start=1
    ):
        success = False
        last_error = None

        for attempt in range(
            1,
            4
        ):
            print(
                f"[{number}/{len(files)}] "
                f"Extracting {file.name} "
                f"(attempt {attempt}/3) ..."
            )

            try:
                result = agent.extract(
                    str(file)
                )

                receipt = Receipt.model_validate(
                    result.data
                )

                data = receipt.model_dump()

                data["category"] = normalise_category(
                    data["category"]
                )

                data["source_file"] = file.name

                extracted.append(
                    data
                )

                print(
                    f"    Merchant : {data['merchant_name']}"
                )

                print(
                    f"    Date     : {data['receipt_date']}"
                )

                print(
                    f"    Category : {data['category']}"
                )

                print(
                    f"    Amount   : "
                    f"{data['currency']} "
                    f"{data['total_hkd']}"
                )

                if data.get(
                    "confidence_note"
                ):
                    print(
                        f"    NOTE     : "
                        f"{data['confidence_note']}"
                    )

                print()

                success = True
                break

            except Exception as error:
                last_error = error

                print(
                    f"    Attempt {attempt} failed: "
                    f"{error}"
                )

        if not success:
            failed.append(
                (
                    file.name,
                    last_error
                )
            )

    if failed:
        names = "\n".join(
            f"• {name}: {error}"
            for name, error in failed
        )

        raise SystemExit(
            f"\nERROR: {len(failed)} of {len(files)} "
            f"receipt(s) could not be extracted "
            f"after 3 attempts.\n\n"
            f"{names}\n\n"
            f"No filled workbook was created because "
            f"every receipt must be included."
        )

    if len(extracted) != len(files):
        raise SystemExit(
            f"\nERROR: Found {len(files)} receipt files "
            f"but extracted only {len(extracted)}. "
            f"No filled workbook was created."
        )

    return extracted


def sort_receipts_chronologically(extracted):
    indexed = list(
        enumerate(
            extracted
        )
    )

    def sort_key(item):
        original_index, receipt = item

        raw_date = str(
            receipt.get(
                "receipt_date",
                ""
            )
        ).strip()

        try:
            parsed_date = datetime.strptime(
                raw_date,
                "%d/%m/%Y"
            )

            return (
                0,
                parsed_date,
                original_index
            )

        except (
            ValueError,
            TypeError
        ):
            return (
                1,
                datetime.max,
                original_index
            )

    sorted_receipts = [
        receipt
        for _, receipt in sorted(
            indexed,
            key=sort_key
        )
    ]

    print(
        "\nReceipts sorted chronologically "
        "(earliest to latest)."
    )

    for item_number, receipt in enumerate(
        sorted_receipts,
        start=1
    ):
        print(
            f"    Item {item_number}: "
            f"{receipt.get('receipt_date', '')} - "
            f"{receipt.get('merchant_name', '')}"
        )

    return sorted_receipts


def get_unique_categories(extracted):
    categories = []

    for receipt in extracted:
        category = normalise_category(
            receipt.get(
                "category"
            )
        )

        if category not in categories:
            categories.append(
                category
            )

    return categories


def snapshot_row(
    ws,
    row,
    max_column=27
):
    cells = []

    for column in range(
        1,
        max_column + 1
    ):
        cell = ws.cell(
            row=row,
            column=column
        )

        cells.append(
            {
                "value": cell.value,
                "style": copy(
                    cell._style
                ),
                "font": copy(
                    cell.font
                ),
                "fill": copy(
                    cell.fill
                ),
                "border": copy(
                    cell.border
                ),
                "alignment": copy(
                    cell.alignment
                ),
                "number_format": cell.number_format,
                "protection": copy(
                    cell.protection
                ),
            }
        )

    dimension = ws.row_dimensions[
        row
    ]

    return {
        "cells": cells,
        "height": dimension.height,
        "hidden": dimension.hidden,
        "outline_level": dimension.outlineLevel,
        "collapsed": dimension.collapsed,
    }


def restore_row(
    ws,
    row,
    snapshot,
    copy_values=True
):
    for column, saved in enumerate(
        snapshot["cells"],
        start=1
    ):
        target = ws.cell(
            row=row,
            column=column
        )

        target._style = copy(
            saved["style"]
        )

        target.font = copy(
            saved["font"]
        )

        target.fill = copy(
            saved["fill"]
        )

        target.border = copy(
            saved["border"]
        )

        target.alignment = copy(
            saved["alignment"]
        )

        target.number_format = saved[
            "number_format"
        ]

        target.protection = copy(
            saved["protection"]
        )

        if copy_values:
            target.value = saved[
                "value"
            ]
        else:
            target.value = None

    dimension = ws.row_dimensions[
        row
    ]

    dimension.height = snapshot[
        "height"
    ]

    dimension.hidden = snapshot[
        "hidden"
    ]

    dimension.outlineLevel = snapshot[
        "outline_level"
    ]

    dimension.collapsed = snapshot[
        "collapsed"
    ]


def snapshot_merges(
    ws,
    first_row,
    last_row
):
    merges = []

    for merged in list(
        ws.merged_cells.ranges
    ):
        min_col, min_row, max_col, max_row = range_boundaries(
            str(
                merged
            )
        )

        if (
            min_row >= first_row
            and max_row <= last_row
        ):
            merges.append(
                (
                    min_col,
                    min_row,
                    max_col,
                    max_row
                )
            )

    return merges


def recreate_shifted_merges(
    ws,
    merges,
    source_first_row,
    destination_first_row
):
    offset = (
        destination_first_row
        - source_first_row
    )

    for (
        min_col,
        min_row,
        max_col,
        max_row
    ) in merges:
        new_min_row = (
            min_row
            + offset
        )

        new_max_row = (
            max_row
            + offset
        )

        start = ws.cell(
            row=new_min_row,
            column=min_col
        ).coordinate

        end = ws.cell(
            row=new_max_row,
            column=max_col
        ).coordinate

        ws.merge_cells(
            f"{start}:{end}"
        )


def wrap_description(
    ws,
    row,
    description
):
    cell = ws[
        f"{DESCRIPTION_COLUMN}{row}"
    ]

    current_width = (
        ws.column_dimensions[
            DESCRIPTION_COLUMN
        ].width
        or 30
    )

    characters_per_line = max(
        15,
        int(
            current_width
            * 0.80
        )
    )

    lines = []

    for paragraph in str(
        description
    ).splitlines():
        paragraph = paragraph.strip()

        if not paragraph:
            lines.append("")
            continue

        wrapped_lines = textwrap.wrap(
            paragraph,
            width=characters_per_line,
            break_long_words=False,
            break_on_hyphens=False,
            replace_whitespace=True,
            drop_whitespace=True,
        )

        if wrapped_lines:
            lines.extend(
                wrapped_lines
            )
        else:
            lines.append(
                paragraph
            )

    if not lines:
        lines = [""]

    cell.value = "\n".join(
        lines
    )

    old_alignment = copy(
        cell.alignment
    )

    cell.alignment = Alignment(
        horizontal=old_alignment.horizontal,
        vertical="top",
        text_rotation=old_alignment.text_rotation,
        wrap_text=True,
        shrink_to_fit=False,
        indent=old_alignment.indent,
        relativeIndent=old_alignment.relativeIndent,
        justifyLastLine=old_alignment.justifyLastLine,
        readingOrder=old_alignment.readingOrder,
    )

    line_count = max(
        1,
        len(
            lines
        )
    )

    required_height = (
        line_count
        * 18
        + 4
    )

    ws.row_dimensions[
        row
    ].height = required_height


def create_category_dropdown(
    wb,
    ws,
    first_receipt_row,
    last_receipt_row
):
    if DROPDOWN_SHEET not in wb.sheetnames:
        raise KeyError(
            f"Worksheet '{DROPDOWN_SHEET}' "
            f"was not found."
        )

    range_name = "ExpenseCategories"

    try:
        if range_name in wb.defined_names:
            del wb.defined_names[
                range_name
            ]

    except Exception:
        pass

    category_range = DefinedName(
        range_name,
        attr_text="'Dropdown'!$A$1:$A$36"
    )

    try:
        wb.defined_names.add(
            category_range
        )

    except AttributeError:
        wb.defined_names.append(
            category_range
        )

    for validation in list(
        ws.data_validations.dataValidation
    ):
        if validation.type == "list":
            try:
                ws.data_validations.dataValidation.remove(
                    validation
                )

            except ValueError:
                pass

    dropdown = DataValidation(
        type="list",
        formula1="=ExpenseCategories",
        allow_blank=True
    )

    dropdown.showErrorMessage = True
    dropdown.errorTitle = "Invalid Category"
    dropdown.error = (
        "Please select a category "
        "from the dropdown list."
    )
    dropdown.showInputMessage = False

    ws.add_data_validation(
        dropdown
    )

    dropdown.add(
        f"{CATEGORY_COLUMN}"
        f"{first_receipt_row}:"
        f"{CATEGORY_COLUMN}"
        f"{last_receipt_row}"
    )


def build_dynamic_form(
    wb,
    ws,
    extracted
):
    receipt_count = max(
        1,
        len(
            extracted
        )
    )

    unique_categories = get_unique_categories(
        extracted
    )

    category_count = max(
        1,
        len(
            unique_categories
        )
    )

    receipt_template = snapshot_row(
        ws,
        RECEIPT_TEMPLATE_ROW
    )

    summary_header_template = snapshot_row(
        ws,
        SUMMARY_HEADER_TEMPLATE_ROW
    )

    category_template = snapshot_row(
        ws,
        CATEGORY_TEMPLATE_ROW
    )

    grand_total_template = snapshot_row(
        ws,
        GRAND_TOTAL_TEMPLATE_ROW
    )

    footer_snapshots = []

    for row in range(
        FOOTER_FIRST_ROW,
        FOOTER_LAST_ROW + 1
    ):
        footer_snapshots.append(
            snapshot_row(
                ws,
                row
            )
        )

    footer_merges = snapshot_merges(
        ws,
        FOOTER_FIRST_ROW,
        FOOTER_LAST_ROW
    )

    master_row_breaks = [
        b.id
        for b in ws.row_breaks.brk
    ]

    image_states = []

    for image in ws._images:
        image_states.append(
            {
                "width": image.width,
                "height": image.height,
                "anchor": deepcopy(
                    image.anchor
                ),
            }
        )

    original_widths = {}

    for column in (
        "A",
        "B",
        "C",
        "D",
        "E",
        "F",
        "G",
        "H",
        "I",
    ):
        original_widths[
            column
        ] = ws.column_dimensions[
            column
        ].width

    first_receipt_row = RECEIPT_TEMPLATE_ROW

    last_receipt_row = (
        first_receipt_row
        + receipt_count
        - 1
    )

    summary_header_row = (
        last_receipt_row
        + 1
    )

    first_category_row = (
        summary_header_row
        + 1
    )

    last_category_row = (
        first_category_row
        + category_count
        - 1
    )

    grand_total_row = (
        last_category_row
        + 1
    )

    footer_first_destination = (
        grand_total_row
        + 1
    )

    for merged in list(
        ws.merged_cells.ranges
    ):
        _, min_row, _, _ = range_boundaries(
            str(
                merged
            )
        )

        if min_row >= RECEIPT_TEMPLATE_ROW:
            ws.unmerge_cells(
                str(
                    merged
                )
            )

    extra_rows = (
        receipt_count
        - 1
        + category_count
        - 1
    )

    if extra_rows > 0:
        ws.insert_rows(
            FOOTER_FIRST_ROW,
            amount=extra_rows
        )

    ws.row_breaks.brk = []

    for break_id in master_row_breaks:
        shifted_break_id = (
            break_id + extra_rows
            if break_id >= FOOTER_FIRST_ROW
            else break_id
        )

        ws.row_breaks.append(
            Break(
                id=shifted_break_id
            )
        )

    final_footer_last_row = (
        footer_first_destination
        + len(
            footer_snapshots
        )
        - 1
    )

    for row in range(
        RECEIPT_TEMPLATE_ROW,
        final_footer_last_row + 1
    ):
        for column in range(
            1,
            28
        ):
            ws.cell(
                row=row,
                column=column
            ).value = None

    for row in range(
        first_receipt_row,
        last_receipt_row + 1
    ):
        restore_row(
            ws,
            row,
            receipt_template,
            copy_values=False
        )

        ws.merge_cells(
            f"E{row}:F{row}"
        )

        ws.merge_cells(
            f"H{row}:I{row}"
        )

        ws[
            f"A{row}"
        ] = (
            row
            - first_receipt_row
            + 1
        )

        ws[
            f"E{row}"
        ] = None

    restore_row(
        ws,
        summary_header_row,
        summary_header_template,
        copy_values=False
    )

    ws.merge_cells(
        f"A{summary_header_row}:"
        f"D{summary_header_row}"
    )

    ws.merge_cells(
        f"E{summary_header_row}:"
        f"I{summary_header_row}"
    )

    ws[
        f"A{summary_header_row}"
    ] = "Category / Categories"

    ws[
        f"E{summary_header_row}"
    ] = "Category Total (HKD)"

    for row in range(
        first_category_row,
        last_category_row + 1
    ):
        restore_row(
            ws,
            row,
            category_template,
            copy_values=False
        )

        ws.merge_cells(
            f"A{row}:D{row}"
        )

        ws.merge_cells(
            f"E{row}:I{row}"
        )

        position = (
            row
            - first_category_row
            + 1
        )

        ws[
            f"A{row}"
        ] = (
            '=IFERROR('
            'INDEX('
            '_xlfn.UNIQUE('
            '_xlfn._xlws.FILTER('
            f'$G${first_receipt_row}:'
            f'$G${last_receipt_row},'
            f'($G${first_receipt_row}:'
            f'$G${last_receipt_row}<>"")'
            '*'
            f'($H${first_receipt_row}:'
            f'$H${last_receipt_row}<>0)'
            ')'
            '),'
            f'{position}'
            '),'
            '""'
            ')'
        )

        ws[
            f"E{row}"
        ] = (
            f'=IF('
            f'A{row}="",'
            f'"",'
            f'IF('
            f'SUMIF('
            f'$G${first_receipt_row}:'
            f'$G${last_receipt_row},'
            f'A{row},'
            f'$H${first_receipt_row}:'
            f'$H${last_receipt_row}'
            f')=0,'
            f'"",'
            f'SUMIF('
            f'$G${first_receipt_row}:'
            f'$G${last_receipt_row},'
            f'A{row},'
            f'$H${first_receipt_row}:'
            f'$H${last_receipt_row}'
            f')'
            f')'
            f')'
        )

    restore_row(
        ws,
        grand_total_row,
        grand_total_template,
        copy_values=False
    )

    ws.merge_cells(
        f"A{grand_total_row}:"
        f"D{grand_total_row}"
    )

    ws.merge_cells(
        f"E{grand_total_row}:"
        f"I{grand_total_row}"
    )

    ws[
        f"A{grand_total_row}"
    ] = "Grand Total (HKD)"

    ws[
        f"E{grand_total_row}"
    ] = (
        f'=SUM('
        f'$H${first_receipt_row}:'
        f'$H${last_receipt_row}'
        f')'
    )

    for index, footer_snapshot in enumerate(
        footer_snapshots
    ):
        destination_row = (
            footer_first_destination
            + index
        )

        restore_row(
            ws,
            destination_row,
            footer_snapshot,
            copy_values=True
        )

    recreate_shifted_merges(
        ws,
        footer_merges,
        FOOTER_FIRST_ROW,
        footer_first_destination
    )

    create_category_dropdown(
        wb,
        ws,
        first_receipt_row,
        last_receipt_row
    )

    for column, width in original_widths.items():
        ws.column_dimensions[
            column
        ].width = width

    for image, saved in zip(
        ws._images,
        image_states
    ):
        image.width = saved[
            "width"
        ]

        image.height = saved[
            "height"
        ]

        image.anchor = deepcopy(
            saved[
                "anchor"
            ]
        )

    return {
        "receipt_count": receipt_count,
        "unique_category_count": len(
            unique_categories
        ),
        "first_receipt_row": first_receipt_row,
        "last_receipt_row": last_receipt_row,
        "summary_header_row": summary_header_row,
        "first_category_row": first_category_row,
        "last_category_row": last_category_row,
        "grand_total_row": grand_total_row,
        "footer_first_row": footer_first_destination,
    }


def write_receipts(
    ws,
    extracted,
    layout
):
    first_receipt_row = layout[
        "first_receipt_row"
    ]

    for index, receipt in enumerate(
        extracted
    ):
        row = (
            first_receipt_row
            + index
        )

        ws[
            f"A{row}"
        ] = index + 1

        ws[
            f"B{row}"
        ] = receipt[
            "receipt_date"
        ]

        ws[
            f"C{row}"
        ] = receipt[
            "merchant_name"
        ]

        wrap_description(
            ws,
            row,
            receipt[
                "description"
            ]
        )

        ws[
            f"E{row}"
        ] = None

        category = normalise_category(
            receipt[
                "category"
            ]
        )

        ws[
            f"G{row}"
        ] = category

        currency = str(
            receipt[
                "currency"
            ]
        ).upper()

        if currency == "HKD":
            ws[
                f"H{row}"
            ] = receipt[
                "total_hkd"
            ]

        else:
            ws[
                f"H{row}"
            ] = None

            print(
                "\n[FOREIGN CURRENCY]"
            )

            print(
                f"Receipt: "
                f"{receipt['source_file']}"
            )

            print(
                f"Original amount: "
                f"{currency} "
                f"{receipt['total_hkd']}"
            )

            print(
                f"Please manually enter "
                f"the converted HKD amount "
                f"in H{row}."
            )


def prepare_receipt_images(
    source_file,
    render_folder
):
    source_path = (
        RECEIPTS_FOLDER
        / source_file
    )

    suffix = source_path.suffix.lower()

    if suffix in (
        ".jpg",
        ".jpeg",
        ".png",
    ):
        return [
            source_path
        ]

    if suffix == ".pdf":
        try:
            import fitz

        except ImportError:
            raise RuntimeError(
                "PDF receipt embedding requires "
                "PyMuPDF. Install it with: "
                "pip install pymupdf"
            )

        render_folder.mkdir(
            parents=True,
            exist_ok=True
        )

        document = fitz.open(
            str(
                source_path
            )
        )

        rendered_pages = []

        try:
            for page_number in range(
                len(
                    document
                )
            ):
                page = document[
                    page_number
                ]

                pixmap = page.get_pixmap(
                    matrix=fitz.Matrix(
                        2.0,
                        2.0
                    ),
                    alpha=False
                )

                output_path = (
                    render_folder
                    / (
                        f"{source_path.stem}"
                        f"_page_"
                        f"{page_number + 1}.png"
                    )
                )

                pixmap.save(
                    str(
                        output_path
                    )
                )

                rendered_pages.append(
                    output_path
                )

        finally:
            document.close()

        return rendered_pages

    raise ValueError(
        f"Unsupported receipt image type: "
        f"{source_path.name}"
    )


def make_receipt_upload_image(
    receipt,
    item_number,
    render_folder
):
    image_paths = prepare_receipt_images(
        receipt[
            "source_file"
        ],
        render_folder
    )

    opened = [
        PILImage.open(
            path
        ).convert(
            "RGB"
        )
        for path in image_paths
    ]

    target_width = max(
        image.width
        for image in opened
    )

    resized = []

    for image in opened:
        if image.width != target_width:
            image = image.resize(
                (
                    target_width,
                    max(
                        1,
                        round(
                            image.height
                            * target_width
                            / image.width
                        )
                    ),
                ),
                PILImage.Resampling.LANCZOS
            )

        resized.append(
            image
        )

    gap = max(
        6,
        target_width // 100
    )

    combined = PILImage.new(
        "RGB",
        (
            target_width,
            sum(
                image.height
                for image in resized
            )
            + gap
            * max(
                0,
                len(
                    resized
                )
                - 1
            ),
        ),
        "white"
    )

    y = 0

    for image in resized:
        combined.paste(
            image,
            (
                0,
                y
            )
        )

        y += (
            image.height
            + gap
        )

    render_folder.mkdir(
        parents=True,
        exist_ok=True
    )

    output_path = (
        render_folder
        / f"item_{item_number:03d}.png"
    )

    combined.save(
        output_path,
        "PNG"
    )

    for image in opened:
        image.close()

    combined.close()

    return output_path


def column_width_pixels(
    ws,
    column
):
    return max(
        1,
        int(
            (
                ws.column_dimensions[
                    column
                ].width
                or 8.43
            )
            * 7
            + 5
        )
    )


def pixel_anchor(
    ws,
    x_pixels,
    row,
    width_pixels,
    height_pixels
):
    remaining = x_pixels
    col_index = 0

    for index, column in enumerate(
        "ABCDEFGHI"
    ):
        pixels = column_width_pixels(
            ws,
            column
        )

        if remaining < pixels:
            col_index = index
            break

        remaining -= pixels

        col_index = min(
            index + 1,
            8
        )

    return OneCellAnchor(
        _from=AnchorMarker(
            col=col_index,
            colOff=pixels_to_EMU(
                max(
                    0,
                    int(
                        remaining
                    )
                )
            ),
            row=row - 1,
            rowOff=0
        ),
        ext=XDRPositiveSize2D(
            pixels_to_EMU(
                int(
                    width_pixels
                )
            ),
            pixels_to_EMU(
                int(
                    height_pixels
                )
            )
        )
    )


def closest_column_to_x(
    ws,
    x
):
    running = 0
    best_col = 1
    best_distance = float(
        "inf"
    )

    for index, column in enumerate(
        "ABCDEFGHI",
        start=1
    ):
        width = column_width_pixels(
            ws,
            column
        )

        center = (
            running
            + width / 2
        )

        distance = abs(
            center
            - x
        )

        if distance < best_distance:
            best_distance = distance
            best_col = index

        running += width

    return best_col


def add_receipt_uploads_section(
    ws,
    extracted,
    layout
):
    banner_row = (
        layout[
            "footer_first_row"
        ]
        + (
            24
            - FOOTER_FIRST_ROW
        )
    )

    section_row = (
        banner_row
        + 1
    )

    ws.sheet_view.view = "pageBreakPreview"

    ws.sheet_properties.pageSetUpPr.fitToPage = True

    ws.page_setup.orientation = "landscape"

    ws.page_setup.paperSize = ws.PAPERSIZE_A4

    ws.page_setup.fitToWidth = 1

    ws.page_setup.fitToHeight = 0

    render_folder = (
        BASE_DIR
        / ".receipt_render_cache"
    )

    total_width = sum(
        column_width_pixels(
            ws,
            c
        )
        for c in "ABCDEFGHI"
    )

    minimum_readable_width = 175
    gutter = 12

    columns_per_page = max(
        1,
        min(
            4,
            int(
                (
                    total_width
                    + gutter
                )
                // (
                    minimum_readable_width
                    + gutter
                )
            )
        )
    )

    slot_width = (
        total_width
        / columns_per_page
    )

    image_width = max(
        minimum_readable_width,
        int(
            slot_width
            - gutter
        )
    )

    usable_page_height = 690

    item_height = 42
    vertical_gap = 10

    receipt_data = []

    for index, receipt in enumerate(
        extracted,
        start=1
    ):
        path = make_receipt_upload_image(
            receipt,
            index,
            render_folder
        )

        with PILImage.open(
            path
        ) as probe:
            natural_width, natural_height = probe.size

        scale = min(
            image_width
            / natural_width,
            1.0
        )

        rendered_width = max(
            1,
            int(
                natural_width
                * scale
            )
        )

        rendered_height = max(
            1,
            int(
                natural_height
                * scale
            )
        )

        receipt_data.append(
            (
                index,
                receipt,
                path,
                rendered_width,
                rendered_height,
            )
        )

    pages = []
    current_page = []
    current_height = 0
    current_row_group = []

    for data in receipt_data:
        current_row_group.append(
            data
        )

        if len(
            current_row_group
        ) == columns_per_page:
            required_height = (
                item_height
                + max(
                    item[4]
                    for item in current_row_group
                )
                + vertical_gap
            )

            if (
                current_page
                and current_height
                + required_height
                > usable_page_height
            ):
                pages.append(
                    current_page
                )

                current_page = []
                current_height = 0

            current_page.append(
                current_row_group
            )

            current_height += required_height

            current_row_group = []

    if current_row_group:
        required_height = (
            item_height
            + max(
                item[4]
                for item in current_row_group
            )
            + vertical_gap
        )

        if (
            current_page
            and current_height
            + required_height
            > usable_page_height
        ):
            pages.append(
                current_page
            )

            current_page = []
            current_height = 0

        current_page.append(
            current_row_group
        )

        current_height += required_height

    if current_page:
        pages.append(
            current_page
        )

    current_row = section_row
    last_upload_row = banner_row

    base_font = copy(
        ws[
            f"A{banner_row}"
        ].font
    )

    base_font.sz = 14
    base_font.bold = True

    for page_index, page in enumerate(
        pages
    ):
        if page_index > 0:
            ws.row_breaks.append(
                Break(
                    id=current_row - 1
                )
            )

        for row_group in page:
            label_row = current_row

            ws.row_dimensions[
                label_row
            ].height = 31.5

            image_row = (
                label_row
                + 1
            )

            tallest = max(
                item[4]
                for item in row_group
            )

            image_rows = max(
                1,
                math.ceil(
                    (
                        tallest
                        + vertical_gap
                    )
                    / 20
                )
            )

            for row in range(
                image_row,
                image_row + image_rows
            ):
                ws.row_dimensions[
                    row
                ].height = 15

            used_label_columns = set()

            for slot, data in enumerate(
                row_group
            ):
                (
                    item_number,
                    receipt,
                    path,
                    rendered_width,
                    rendered_height,
                ) = data

                slot_left = (
                    slot
                    * slot_width
                )

                center_x = (
                    slot_left
                    + slot_width / 2
                )

                label_col = closest_column_to_x(
                    ws,
                    center_x
                )

                if label_col in used_label_columns:
                    candidates = sorted(
                        range(
                            1,
                            10
                        ),
                        key=lambda column:
                        abs(
                            column
                            - label_col
                        )
                    )

                    label_col = next(
                        column
                        for column in candidates
                        if column not in used_label_columns
                    )

                used_label_columns.add(
                    label_col
                )

                item_cell = ws.cell(
                    label_row,
                    label_col
                )

                item_cell.value = (
                    f"Item {item_number}"
                )

                item_cell.font = copy(
                    base_font
                )

                item_cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center"
                )

                item_cell.border = Border(
                    left=Side(
                        style="thin"
                    ),
                    right=Side(
                        style="thin"
                    ),
                    top=Side(
                        style="thin"
                    ),
                    bottom=Side(
                        style="thin"
                    ),
                )

                image = XLImage(
                    str(
                        path
                    )
                )

                scale = min(
                    image_width
                    / image.width,
                    1.0
                )

                image.width = max(
                    1,
                    int(
                        image.width
                        * scale
                    )
                )

                image.height = max(
                    1,
                    int(
                        image.height
                        * scale
                    )
                )

                label_left = sum(
                    column_width_pixels(
                        ws,
                        column
                    )
                    for column in "ABCDEFGHI"[
                        :label_col - 1
                    ]
                )

                label_center = (
                    label_left
                    + column_width_pixels(
                        ws,
                        "ABCDEFGHI"[
                            label_col - 1
                        ]
                    )
                    / 2
                )

                image_x = int(
                    label_center
                    - image.width / 2
                )

                image.anchor = pixel_anchor(
                    ws,
                    image_x,
                    image_row,
                    image.width,
                    image.height
                )

                ws.add_image(
                    image
                )

            current_row = (
                image_row
                + image_rows
            )

        last_upload_row = (
            current_row
            - 1
        )

    ws.print_area = (
        f"A1:I{last_upload_row}"
    )

    print(
        f"\nReceipt Uploads banner row: "
        f"{banner_row}"
    )

    print(
        f"Receipt images begin at row: "
        f"{section_row}"
    )

    print(
        f"Receipt Uploads ends at row: "
        f"{last_upload_row}"
    )

    print(
        f"Receipt images placed: "
        f"{len(receipt_data)}"
    )

    print(
        f"Receipt upload pages created: "
        f"{len(pages)}"
    )

    print(
        "The page break already present "
        "in the master workbook was preserved."
    )

    print(
        "Automatic page breaks are added "
        "only after a receipt-upload page "
        "has been filled."
    )

    return {
        "receipt_uploads_banner_row": banner_row,
        "receipt_uploads_first_row": section_row,
        "receipt_uploads_last_row": last_upload_row,
        "receipt_render_folder": render_folder,
    }


def verify_generated_workbook():
    try:
        verification_wb = load_workbook(
            OUTPUT_FILE,
            data_only=False
        )

        if SHEET_NAME not in verification_wb.sheetnames:
            raise RuntimeError(
                "Form worksheet missing."
            )

        verification_ws = verification_wb[
            SHEET_NAME
        ]

        _ = list(
            verification_ws.merged_cells.ranges
        )

        _ = list(
            verification_ws.data_validations.dataValidation
        )

        _ = verification_ws.max_row

        verification_wb.close()

    except Exception as error:
        raise SystemExit(
            "\nERROR: Generated workbook failed "
            "the integrity check.\n\n"
            f"{error}\n\n"
            "The master workbook was not modified."
        )


def process_receipts():
    check_master_file()

    check_receipts_folder()

    receipt_files = get_receipt_files()

    extracted = extract_receipts(
        receipt_files
    )

    if len(extracted) != len(receipt_files):
        raise SystemExit(
            f"\nERROR: Found "
            f"{len(receipt_files)} receipt files "
            f"but only extracted "
            f"{len(extracted)}."
        )

    extracted = sort_receipts_chronologically(
        extracted
    )

    try:
        shutil.copy2(
            MASTER_FILE,
            OUTPUT_FILE
        )

    except PermissionError:
        raise SystemExit(
            "\nERROR: Expense_Claim_filled.xlsx "
            "is currently open in Excel.\n\n"
            "Close it completely and run again."
        )

    wb = load_workbook(
        OUTPUT_FILE
    )

    if SHEET_NAME not in wb.sheetnames:
        raise SystemExit(
            f"\nERROR: Worksheet "
            f"'{SHEET_NAME}' was not found."
        )

    ws = wb[
        SHEET_NAME
    ]

    layout = build_dynamic_form(
        wb,
        ws,
        extracted
    )

    write_receipts(
        ws,
        extracted,
        layout
    )

    upload_layout = add_receipt_uploads_section(
        ws,
        extracted,
        layout
    )

    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.calculation.calcMode = "auto"

    except Exception:
        pass

    try:
        wb.save(
            OUTPUT_FILE
        )

        wb.close()

    except PermissionError:
        raise SystemExit(
            "\nERROR: Expense_Claim_filled.xlsx "
            "is currently open in Excel.\n\n"
            "Close it completely and run again."
        )

    verify_generated_workbook()

    render_folder = upload_layout.get(
        "receipt_render_folder"
    )

    if (
        render_folder
        and render_folder.exists()
    ):
        shutil.rmtree(
            render_folder,
            ignore_errors=True
        )

    print(
        "\n========================================"
    )

    print(
        "EXPENSE CLAIM CREATED"
    )

    print(
        "========================================"
    )

    print(
        f"\nReceipt files found: "
        f"{len(receipt_files)}"
    )

    print(
        f"Receipts successfully processed: "
        f"{layout['receipt_count']}"
    )

    print(
        f"Receipt images inserted: "
        f"{len(extracted)}"
    )

    print(
        f"Receipt rows: "
        f"{layout['first_receipt_row']}"
        f"-"
        f"{layout['last_receipt_row']}"
    )

    print(
        f"Unique categories: "
        f"{layout['unique_category_count']}"
    )

    print(
        f"Category rows: "
        f"{layout['first_category_row']}"
        f"-"
        f"{layout['last_category_row']}"
    )

    print(
        f"Grand Total row: "
        f"{layout['grand_total_row']}"
    )

    print(
        "\nAll receipts were successfully "
        "included."
    )

    print(
        "Receipts were ordered "
        "chronologically."
    )

    print(
        "The master workbook page break "
        "was preserved."
    )

    print(
        "Automatic page breaks occur only "
        "inside the receipt-image section "
        "after a page has been filled."
    )

    print(
        "\nOpen Expense_Claim_filled.xlsx "
        "and manually check ALL information "
        "before printing/submitting."
    )

    print(
        "========================================\n"
    )


def reset_expense_claim():
    check_master_file()

    try:
        shutil.copy2(
            MASTER_FILE,
            OUTPUT_FILE
        )

    except PermissionError:
        raise SystemExit(
            "\nERROR: Expense_Claim_filled.xlsx "
            "is currently open in Excel.\n\n"
            "Close Excel completely "
            "and run Reset again."
        )

    print(
        "\n========================================"
    )

    print(
        "EXPENSE CLAIM RESET"
    )

    print(
        "========================================"
    )

    print(
        "\nThe exact master workbook "
        "has been restored."
    )

    print(
        "========================================\n"
    )


def main_menu():
    print(
        "\n========================================"
    )

    print(
        "        EXPENSE CLAIM TOOL"
    )

    print(
        "========================================"
    )

    print(
        "\n1 - Process receipts"
    )

    print(
        "2 - Reset expense claim"
    )

    print(
        "3 - Exit"
    )

    choice = input(
        "\nChoose 1, 2 or 3: "
    ).strip()

    if choice == "1":
        process_receipts()

    elif choice == "2":
        reset_expense_claim()

    elif choice == "3":

        input(
            "\nPress Enter to exit...\n"
        )

    else:
        print(
            "\nInvalid selection."
        )

        print(
            "Run python llamacode.py again "
            "and choose 1, 2 or 3.\n"
        )


if __name__ == "__main__":
    main_menu()